import json
import sys
from datetime import datetime
import hashlib

from openpyxl import load_workbook


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")


HEADER_TIME = "交易时间"
HEADER_TRADE_TYPE = "交易类型"
HEADER_PRODUCT = "商品"
HEADER_IN_OUT = "收/支"
HEADER_AMOUNT = "金额(元)"
HEADER_NOTE = "备注"


def find_header_map(sheet):
    for row in sheet.iter_rows(values_only=True):
        values = [str(cell).strip() if cell is not None else "" for cell in row]
        if HEADER_TIME in values:
            header_map = {}
            for index, value in enumerate(values):
                if value:
                    header_map[value] = index
            return header_map
    raise ValueError("未找到微信账单表头")


def get_cell(row, header_map, header_name):
    index = header_map.get(header_name)
    if index is None or index >= len(row):
        return ""
    return row[index]


def normalize_type(in_out_text, trade_type_text):
    in_out = str(in_out_text or "").strip()
    trade_type = str(trade_type_text or "").strip()

    if "收入" in in_out:
        return "收入"
    if "支出" in in_out:
        return "支出"

    income_keywords = ["二维码收款", "收款", "退款", "退款到账", "转入"]
    expense_keywords = ["转账", "扫码付款", "商户消费", "充值", "支付"]

    if any(keyword in trade_type for keyword in income_keywords):
        return "收入"
    if any(keyword in trade_type for keyword in expense_keywords):
        return "支出"
    return ""


def normalize_category(trade_type_text, product_text, note_text):
    source = " ".join(
        [
            str(trade_type_text or "").strip(),
            str(product_text or "").strip(),
            str(note_text or "").strip(),
        ]
    )

    if any(keyword in source for keyword in ["餐", "外卖", "奶茶", "咖啡", "饭", "小吃"]):
        return "餐饮"
    if any(keyword in source for keyword in ["公交", "地铁", "打车", "车费", "出行"]):
        return "交通"
    if any(keyword in source for keyword in ["超市", "购物", "淘宝", "京东", "商品"]):
        return "购物"
    if any(keyword in source for keyword in ["工资", "薪资", "兼职", "报酬"]):
        return "工资"
    if any(keyword in source for keyword in ["学", "考试", "课程", "书", "校园"]):
        return "学习"
    if any(keyword in source for keyword in ["视频", "会员", "游戏", "娱乐"]):
        return "娱乐"
    if any(keyword in source for keyword in ["房租", "宿舍", "住宿", "电费", "水费"]):
        return "住房"
    if any(keyword in source for keyword in ["医院", "药", "医疗", "诊所"]):
        return "医疗"
    return "其他"


def normalize_bill_date(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    text = str(value or "").strip()
    if not text:
        return ""
    text = text.replace("/", "-")
    text = " ".join(text.split())
    return text


def normalize_note(note_text, product_text, trade_type_text):
    note = str(note_text or "").strip()
    product = str(product_text or "").strip()
    trade_type = str(trade_type_text or "").strip()
    if note and note != "/":
        return note
    if product and product != "/":
        return product
    return trade_type


def build_fingerprint(trade_time, trade_type, product, in_out, amount, note):
    source = "|".join(
        [
            str(trade_time or "").strip(),
            str(trade_type or "").strip(),
            str(product or "").strip(),
            str(in_out or "").strip(),
            str(amount or "").strip(),
            str(note or "").strip(),
        ]
    )
    return hashlib.sha256(source.encode("utf-8")).hexdigest()


def parse_file(file_path):
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header_map = find_header_map(sheet)
    records = []
    header_found = False

    for row in sheet.iter_rows(values_only=True):
        values = [str(cell).strip() if cell is not None else "" for cell in row]
        if not header_found:
            if HEADER_TIME in values:
                header_found = True
            continue

        trade_time = get_cell(row, header_map, HEADER_TIME)
        if trade_time in ("", None):
            continue

        trade_type = get_cell(row, header_map, HEADER_TRADE_TYPE)
        product = get_cell(row, header_map, HEADER_PRODUCT)
        in_out = get_cell(row, header_map, HEADER_IN_OUT)
        amount = get_cell(row, header_map, HEADER_AMOUNT)
        note = get_cell(row, header_map, HEADER_NOTE)

        record_type = normalize_type(in_out, trade_type)
        if not record_type:
            continue

        try:
            amount_value = float(amount)
        except Exception:
            continue

        bill_date = normalize_bill_date(trade_time)
        if not bill_date:
            continue

        records.append(
            {
                "type": record_type,
                "amount": amount_value,
                "category": normalize_category(trade_type, product, note),
                "billDate": bill_date,
                "note": normalize_note(note, product, trade_type),
                "fingerprint": build_fingerprint(trade_time, trade_type, product, in_out, amount, note),
            }
        )

    return records


def main():
    if len(sys.argv) < 2:
        raise SystemExit("missing file path")
    file_path = sys.argv[1]
    records = parse_file(file_path)
    print(json.dumps({"records": records}, ensure_ascii=False))


if __name__ == "__main__":
    main()
